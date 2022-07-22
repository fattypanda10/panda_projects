import os
from typing import List
import re

import pandas as pd
from natsort import natsorted
from openpyxl import load_workbook


def extract_data(file_path: str, sheet_name: str) -> list:
    extracted_data = []
    if sheet_name[:-2] in ('1-2', '3-4-5', '6', '17', '20'):
        extracted_data.extend(pd.Series(get_rating_data(file_path, sheet_name), dtype=object))
    elif sheet_name[:-2] == '13-14':
        extracted_data.extend(get_data_from_ques_13_14(file_path, sheet_name))
    else:
        extracted_data.extend(get_data_from_choice_ques(file_path, sheet_name))
    return extracted_data


def get_sheetnames(filepath: str) -> list:
    wb = load_workbook(filepath, read_only=True)
    return wb.sheetnames


def read_excel(filepath: str, sheetname: str, cols_use: str = None, rows_skip: List[int] = None,
               num_rows: int = None) -> pd.DataFrame:
    dataframe = pd.read_excel(filepath,
                              sheet_name=sheetname,
                              usecols=cols_use,
                              index_col=None,
                              skiprows=rows_skip,
                              nrows=num_rows,
                              keep_default_na=False)
    return dataframe


def get_rating_data(filepath: str, sheetname: str) -> list:
    cols_with_vals, idx_for_drop = [], []
    dataframe = read_excel(filepath,
                           sheetname=sheetname,
                           cols_use='C:G',
                           num_rows=20
                           )
    dataframe = del_empty_rows(dataframe)
    if sheetname[:-2] == '1-2':
        idx_for_drop = [0, 1, 2, 4, 5, 6]
    elif sheetname[:-2] == '3-4-5':
        idx_for_drop = [0, 1, 2, 4, 5, 6, 8]
    elif sheetname[:-2] == '17':
        idx_for_drop = [0, 1, 2]
    dataframe.drop(index=idx_for_drop, inplace=True)
    dataframe.reset_index(inplace=True, drop=True)
    for row in dataframe.itertuples():
        if row[0] in (0, 1):
            cols_with_vals.append([i != '' for i in row[1:]].index(True) + 1)
        else:
            cols_with_vals.append(dataframe.iloc[row[0], 0])
    return cols_with_vals


def del_empty_rows(dataframe: pd.DataFrame):
    for row in dataframe.itertuples():
        # here the index is dependent on the cols_used in read_excel
        if all([i == '' for i in row[1:]]):
            dataframe.drop(index=row[0], inplace=True)
    dataframe.reset_index(inplace=True, drop=True)
    return dataframe


def compute_rows_to_del_from_top(dataframe: pd.DataFrame):
    dataframe_rows = list(dataframe.itertuples())
    if 1 in dataframe.values:
        for ind, val in enumerate(dataframe_rows):
            if not isinstance(val[1], str):
                return ind
    else:
        for ind, val in enumerate(dataframe_rows):
            if val[2]:
                return ind


def del_rows_from_top(dataframe: pd.DataFrame) -> pd.DataFrame:
    if compute_rows_to_del_from_top(dataframe):
        dataframe.drop(index=[i for i in range(compute_rows_to_del_from_top(dataframe))], inplace=True)
        dataframe.reset_index(inplace=True, drop=True)
    return dataframe


def del_unwanted_rows(dataframe: pd.DataFrame) -> pd.DataFrame:
    for i in list(dataframe.itertuples()):
        if not i[1]:
            if not i[3] or (isinstance(i[3], str) and len(i[3]) > 8):
                dataframe.drop(index=i[0], inplace=True)
    dataframe.reset_index(inplace=True, drop=True)
    return dataframe


def del_remaining_unwanted_rows(dataframe: pd.DataFrame) -> pd.DataFrame:
    for i in list(dataframe.itertuples()):
        if re.search(r'\d{2}[\-]..[\-]\d{2}', str(i[3])):
            continue
        if (not i[1] and not i[2]) or (not isinstance(i[1], int)):
            dataframe.drop(index=i[0], inplace=True)
    dataframe.reset_index(inplace=True, drop=True)
    return dataframe


def clean_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """
    1. Removes completely-empty rows
    2. Deletes irrelevant rows from top
    3. Deletes unwanted rows from within the DataFrame
    4. Remove remaining irrelevant rows
    Resets the index after every removal

    """
    return del_remaining_unwanted_rows(del_unwanted_rows(del_rows_from_top(del_empty_rows(dataframe))))


def get_data_from_choice_ques(filepath: str, sheetname: str):
    double_rows_dict = {'7': 22, '8': 21}
    dataframe = read_excel(filepath, sheetname, rows_skip=[2])
    dataframe.replace(['V', 'v', 'P'], 1, inplace=True)
    dataframe = clean_dataframe(dataframe)
    if sheetname[:-2] in ('7', '8'):
        dataframe.drop(index=double_rows_dict[sheetname[:-2]], inplace=True)
        dataframe.reset_index(inplace=True, drop=True)
    dataframe.replace('', 0, inplace=True)
    if len(dataframe.columns) >= 3 and sum([1 if i != object else 0 for i in dataframe.dtypes[1:]]) >= 2:
        dataframe_combined = pd.concat([
            dataframe[i] for i in dataframe.columns[-2:]],
            axis=0,
            ignore_index=True,
        )
        return dataframe_combined.transpose()
    return dataframe[dataframe.columns[-1]]


def get_data_from_ques_13_14(filepath: str, sheetname: str) -> pd.DataFrame:
    dataframe = pd.read_excel(filepath,
                              sheet_name=sheetname,
                              usecols='B:C',
                              skiprows=[i - 1 for i in [2, 3, 9, 10, 11]],
                              index_col=None,
                              keep_default_na=False)
    return dataframe[dataframe.columns[-1]]


def export_to_csv(data: list, filename: str) -> None:
    export_path = os.path.join(os.getcwd(), filename)
    pd.DataFrame(data).transpose().to_csv(export_path, header=False, index=False, mode='a')
